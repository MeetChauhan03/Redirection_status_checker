
import streamlit as st
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import urllib3

# === Configuration ===
st.set_page_config(page_title="Redirect Auditor Pro", layout="wide", page_icon="🔗")
MAX_WORKERS = 20
DEFAULT_TIMEOUT = 10

# === Headers Strategy (The Fix) ===
# Real browser headers to bypass Akamai/AEM bot detection (prevents false 301s)
BROWSER_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Cache-Control': 'no-cache',
    'Pragma': 'no-cache',
}

# Standard python-requests user agent (for comparison/debugging)
BOT_HEADERS = {
    'User-Agent': 'python-requests/2.31.0',
    'Accept-Encoding': 'gzip, deflate',
}

# === HTTP Status Descriptions ===
status_names = {
    200: 'OK', 301: 'Moved Permanently', 302: 'Found', 303: 'See Other',
    307: 'Temporary Redirect', 308: 'Permanent Redirect', 400: 'Bad Request',
    401: 'Unauthorized', 403: 'Forbidden', 404: 'Not Found', 500: 'Internal Server Error'
}

# === Utility Functions ===

def get_server_name(headers):
    akamai_indicators = ["AkamaiGHost", "akamaitechnologies.com", "X-Akamai-Transformed"]
    combined_headers = " | ".join(f"{k}: {v}" for k, v in headers.items())
    
    for marker in akamai_indicators:
        if marker.lower() in combined_headers.lower():
            return "Akamai"
    
    server_headers_priority = ["Server", "X-Powered-By", "Via", "CF-RAY", "X-CDN"]
    for key in server_headers_priority:
        if key in headers:
            return f"{key}: {headers[key]}"
    return "Unknown"

def is_blocked_url(url):
    return "b2b-b" in url.lower()

def check_redirection_chain(url, use_browser_headers=True, verify_ssl=True, timeout=DEFAULT_TIMEOUT):
    if not verify_ssl:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    headers = BROWSER_HEADERS if use_browser_headers else BOT_HEADERS
    visited = set()
    chain = []
    current_url = url.strip()

    try:
        while True:
            if current_url in visited:
                chain.append({
                    'URL': current_url, 'Status': 'Loop detected',
                    'Status Code': 'Loop', 'Server': 'N/A'
                })
                break

            visited.add(current_url)
            
            # CRITICAL: allow_redirects=False is needed to capture the specific 301 vs 302 code
            resp = requests.get(
                current_url, 
                timeout=timeout, 
                allow_redirects=False, 
                headers=headers,
                verify=verify_ssl
            )
            
            status = resp.status_code
            status_text = status_names.get(status, 'Unknown')
            server = get_server_name(resp.headers)
            
            chain.append({
                'URL': current_url, 'Status': status_text,
                'Status Code': status, 'Server': server
            })

            if status in (301, 302, 303, 307, 308):
                redirect_url = resp.headers.get('Location')
                if not redirect_url:
                    break
                
                if redirect_url.startswith('/'):
                    from urllib.parse import urljoin
                    redirect_url = urljoin(current_url, redirect_url)
                current_url = redirect_url
            else:
                break
                
    except requests.exceptions.RequestException as e:
        chain.append({
            'URL': current_url, 'Status': 'Connection Error',
            'Status Code': 'Error', 'Server': 'N/A'
        })
    except Exception as e:
        chain.append({
            'URL': current_url, 'Status': f"Error: {str(e)}",
            'Status Code': 'Error', 'Server': 'N/A'
        })
    
    return chain

def render_markdown_chain(chain):
    if not chain: return "No data"
    lines = []
    for i, step in enumerate(chain):
        icon = "🟢"
        code = step['Status Code']
        if code == 'Loop': icon = "🔄"
        elif code == 'Error': icon = "❌"
        elif isinstance(code, int):
            if 300 <= code < 400: icon = "🟡"
            elif code >= 400: icon = "🔴"
        
        indent = " " * (i * 4)
        arrow = "└─>" if i > 0 else "🚩"
        lines.append(f"{indent} {arrow} {icon} **{code}** : `{step['URL']}`  \n({step['Status']} | Server: {step['Server']})")
    return "  \n".join(lines)

def generate_excel(df_summary, df_tracking):
    output = BytesIO()
    wb = Workbook()
    
    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    # Headers
    headers1 = ["Original URL", "Final URL", "Final Status Code", "Server", "Chain Length"]
    ws1.append(headers1)
    
    # Styling Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="36454F", end_color="36454F", fill_type="solid")
    
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in dataframe_to_rows(df_summary, index=False, header=False):
        ws1.append(r)
        
    # Auto-width roughly
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 25

    # Sheet 2: Detailed Tracking
    ws2 = wb.create_sheet("Detailed Tracking")
    headers2 = ["Original URL", "Step", "Hop URL", "Status Code", "Status Description", "Server"]
    ws2.append(headers2)
    
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for r in dataframe_to_rows(df_tracking, index=False, header=False):
        ws2.append(r)
        # Color coding rows based on status
        status_val = r[3] # Status Code index
        fill_color = None
        if status_val == 200: fill_color = "E2EFDA" # Green
        elif str(status_val).startswith('3'): fill_color = "FFF2CC" # Yellow
        elif str(status_val).startswith('4') or str(status_val).startswith('5'): fill_color = "FCE4D6" # Red
        
        if fill_color:
            for cell in ws2[ws2.max_row]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 25

    wb.save(output)
    output.seek(0)
    return output

# === SIDEBAR UI ===
with st.sidebar:
    st.markdown("### 🛠️ Configuration")
    
    # Input Section
    st.markdown("**1. Input Source**")
    uploaded_file = st.file_uploader("Upload Excel (.xlsx)", type="xlsx")
    if uploaded_file:
        st.success(f"File loaded: {uploaded_file.name}")
        
    text_input = st.text_area("Or paste URLs (one per line):", height=150)
    
    st.divider()
    
    # Advanced Settings
    with st.expander("⚙️ Advanced Settings", expanded=True):
        use_browser_ua = st.toggle("Simulate Real Browser", value=True, 
                                 help="Use Chrome User-Agent headers to fix AEM/Akamai 301 vs 302 detection issues.")
        verify_ssl = st.toggle("Verify SSL Certificates", value=True)
        timeout_val = st.slider("Request Timeout (sec)", 1, 30, DEFAULT_TIMEOUT)
    
    # Action Button
    run_check = st.button("▶ Run Audit", type="primary", use_container_width=True)
    
    # Sample Download
    st.markdown("---")
    sample_df = pd.DataFrame({"Original URL": ["https://example.com", "https://google.com"]})
    sample_csv = sample_df.to_csv(index=False).encode('utf-8')
    st.download_button("📥 Download Sample Template", data=sample_csv, file_name="sample_urls.csv", mime="text/csv", use_container_width=True)

# === MAIN UI ===
st.title("🔗 Redirect Auditor Pro")
st.markdown("""
    **Audit your URL redirection chains with precision.**  
    Features accurate 301/302 detection for AEM/Akamai environments by simulating real browser requests.
""")

# Logic Processing
urls_to_check = []
if uploaded_file:
    try:
        df_in = pd.read_excel(uploaded_file)
        if 'Original URL' in df_in.columns:
            urls_to_check.extend(df_in['Original URL'].dropna().astype(str).tolist())
        else:
            st.error("Uploaded file must contain a column named 'Original URL'")
    except Exception as e:
        st.error(f"Error reading file: {e}")

if text_input:
    urls_to_check.extend([u.strip() for u in text_input.splitlines() if u.strip()])

# Remove duplicates
urls_to_check = list(dict.fromkeys(urls_to_check))

if run_check and urls_to_check:
    # 1. Progress Bar
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    results = []
    
    # 2. Execution
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(check_redirection_chain, url, use_browser_ua, verify_ssl, timeout_val): url for url in urls_to_check}
        total = len(urls_to_check)
        
        for i, future in enumerate(as_completed(futures)):
            url = futures[future]
            chain = future.result()
            results.append((url, chain))
            
            # Update progress
            progress = (i + 1) / total
            progress_bar.progress(progress)
            status_text.text(f"Checking {i+1}/{total}: {url}")
            
    progress_bar.empty()
    status_text.empty()
    
    # 3. Data Processing for Display
    summary_data = []
    detailed_data = []
    
    for orig_url, chain in results:
        final_step = chain[-1]
        summary_data.append({
            "Original URL": orig_url,
            "Final URL": final_step['URL'],
            "Final Status Code": final_step['Status Code'],
            "Server": final_step['Server'],
            "Chain Length": len(chain)
        })
        
        for idx, step in enumerate(chain):
            detailed_data.append({
                "Original URL": orig_url,
                "Step": idx + 1,
                "Hop URL": step['URL'],
                "Status Code": step['Status Code'],
                "Status Description": step['Status'],
                "Server": step['Server']
            })

    df_summary = pd.DataFrame(summary_data)
    df_detailed = pd.DataFrame(detailed_data)

    # 4. Dashboard Metrics
    st.markdown("### 📊 Audit Results")
    m1, m2, m3, m4 = st.columns(4)
    
    success_cnt = df_summary[df_summary['Final Status Code'] == 200].shape[0]
    error_cnt = df_summary[df_summary['Final Status Code'].isin(['Error', 'Loop'])].shape[0]
    # Count URLs that ended in 404
    not_found_cnt = df_summary[df_summary['Final Status Code'] == 404].shape[0]
    
    m1.metric("Total URLs", len(df_summary))
    m2.metric("Success (200 OK)", success_cnt)
    m3.metric("Broken (404)", not_found_cnt)
    m4.metric("Errors/Loops", error_cnt)
    
    # 5. Result Table
    st.dataframe(
        df_summary, 
        use_container_width=True,
        column_config={
            "Final Status Code": st.column_config.NumberColumn("Status", format="%d"),
            "Original URL": st.column_config.LinkColumn("Original"),
            "Final URL": st.column_config.LinkColumn("Final"),
        }
    )
    
    # 6. Detailed Analysis & Download
    col_left, col_right = st.columns([2, 1])
    
    with col_left:
        st.markdown("#### ⛓️ Detailed Chains")
        for orig_url, chain in results:
            with st.expander(f"{orig_url}  →  {chain[-1]['Status Code']}"):
                st.markdown(render_markdown_chain(chain), unsafe_allow_html=True)
                
    with col_right:
        st.info("**Export Report**\n\nIncludes two tabs:\n1. Executive Summary\n2. Detailed Hop-by-Hop Tracking")
        excel_data = generate_excel(df_summary, df_detailed)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name="redirect_audit_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

elif run_check and not urls_to_check:
    st.warning("⚠️ Please upload a file or paste URLs to begin.")